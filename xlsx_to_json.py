# -*- coding: utf-8 -*-
"""
xlsx_to_json.py — 二建备考 App 题库转换器
==========================================
用途: 把 Excel/CSV/JSON 题库文件,转成符合 App DATA schema 的 JSON 文件
依赖: Python 3.x (推荐 3.12),openpyxl 库 (xlsx 读取)

支持题型:
  - recite  (背诵卡:    q/a/sec/sub)
  - quiz    (选择题:    q/choices/answer/analysis/sec)
  - exam    (真题:      同 quiz)
  - case    (案例题:    sec/scenario/question/points)

输出: 同目录下生成 题目类型_yyyymmdd_HHMMSS.json
"""
import os, sys, json, csv, argparse
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print('❌ 缺少 openpyxl 库。请执行: pip install openpyxl')
    sys.exit(1)


# ---------- 核心转换函数 ----------

def parse_answer(ans_str):
    """解析答案字符串: 'A' → 0, 'B,C' → [1,2], 'BD' → [1,3]"""
    if not ans_str:
        return None
    s = str(ans_str).strip().upper().replace(' ', '')
    if ',' in s or ';' in s:
        # 多选
        seps = [',', ';', '、']
        for sep in seps:
            if sep in s:
                s = s.replace(sep, ',')
        chars = [c for c in s.split(',') if c]
        return [_letter_to_idx(c) for c in chars if _letter_to_idx(c) is not None]
    return _letter_to_idx(s)


def _letter_to_idx(c):
    """A→0, B→1, ..., H→7"""
    if not c or len(c) != 1:
        return None
    if 'A' <= c <= 'Z':
        return ord(c) - ord('A')
    return None


def parse_choices(csv_or_list):
    """解析选项列:支持 'A. xxx|B. yyy' 或 5 行单列或 ['A','B','C','D']"""
    if isinstance(csv_or_list, list):
        if not csv_or_list:
            return []
        if len(csv_or_list) == 1 and ('|' in csv_or_list[0] or '\n' in csv_or_list[0]):
            return _split_choices_str(csv_or_list[0])
        return csv_or_list
    return _split_choices_str(str(csv_or_list))


def _split_choices_str(s):
    """按 '|' 或换行符切分选项,自动去 A./B. 前缀"""
    parts = re.split(r'[|\n]', s)
    cleaned = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        # 去 "A. " / "A、 " 前缀
        import re as _re
        p = _re.sub(r'^[A-H][.\s、:：]\s*', '', p)
        cleaned.append(p)
    return cleaned


def cell_val(cell):
    """取单元格值(去除换行/前后空格)"""
    if cell is None:
        return ''
    v = cell.value
    if v is None:
        return ''
    return str(v).strip()


# ---------- 各题型转换器 ----------

def xlsx_to_recite(ws, sec_col=None):
    """背诵卡:列假设 id | q | a | sec | sub(可选)"""
    items = []
    headers = [cell_val(c).lower() for c in ws[1]]
    if 'id' not in headers:
        raise ValueError('背诵表必须包含 id 列')
    q_idx = headers.index('q') if 'q' in headers else None
    a_idx = headers.index('a') if 'a' in headers else None
    sec_idx = headers.index('sec') if 'sec' in headers else None
    sub_idx = headers.index('sub') if 'sub' in headers else None
    if q_idx is None or a_idx is None:
        raise ValueError('背诵表必须包含 q 和 a 列')
    for row in ws.iter_rows(min_row=2, values_only=False):
        rid = cell_val(row[headers.index('id')])
        if not rid:
            continue
        item = {
            'id': rid,
            'q': cell_val(row[q_idx]) if q_idx is not None else '',
            'a': cell_val(row[a_idx]) if a_idx is not None else '',
        }
        if sec_idx is not None:
            sec = cell_val(row[sec_idx])
            if sec:
                item['sec'] = sec
        if sub_idx is not None:
            sub = cell_val(row[sub_idx])
            if sub:
                item['sub'] = sub
        if item['q'] and item['a']:
            items.append(item)
    return items


def xlsx_to_quiz(ws, type_name='quiz'):
    """选择题(quiz/exam 共用):
    列假设 id | q | A | B | C | D | E | answer | analysis | sec
    或: id | q | choices(json/分隔) | answer | analysis | sec
    """
    items = []
    headers = [cell_val(c).lower() for c in ws[1]]
    if 'id' not in headers:
        raise ValueError(f'{type_name} 表必须包含 id 列')
    q_idx = headers.index('q') if 'q' in headers else None
    if q_idx is None:
        raise ValueError(f'{type_name} 表必须包含 q 列')
    # 优先检测 A-H 单列
    choice_cols = [i for i, h in enumerate(headers) if h in 'abcdefgh']
    choices_col = None
    if not choice_cols and 'choices' in headers:
        choices_col = headers.index('choices')
    if not choice_cols and choices_col is None:
        raise ValueError(f'{type_name} 表必须包含 A-H 选项列 或 choices 列')
    ans_idx = headers.index('answer') if 'answer' in headers else None
    if ans_idx is None:
        raise ValueError(f'{type_name} 表必须包含 answer 列')
    ana_idx = headers.index('analysis') if 'analysis' in headers else None
    sec_idx = headers.index('sec') if 'sec' in headers else None
    for row in ws.iter_rows(min_row=2, values_only=False):
        rid = cell_val(row[headers.index('id')])
        if not rid:
            continue
        if choice_cols:
            choices = [cell_val(row[i]) for i in choice_cols if cell_val(row[i])]
        else:
            choices = parse_choices([cell_val(row[choices_col])])
        ans_raw = cell_val(row[ans_idx])
        ans = parse_answer(ans_raw)
        if ans is None or (isinstance(ans, list) and len(ans) == 0):
            print(f'  ⚠ {rid}: 答案 "{ans_raw}" 无法解析,跳过')
            continue
        item = {
            'id': rid,
            'q': cell_val(row[q_idx]),
            'choices': choices,
            'answer': ans,
        }
        if ana_idx is not None:
            ana = cell_val(row[ana_idx])
            if ana:
                item['analysis'] = ana
        if sec_idx is not None:
            sec = cell_val(row[sec_idx])
            if sec:
                item['sec'] = sec
        if item['q'] and item['choices']:
            items.append(item)
    return items


def xlsx_to_case(ws):
    """案例题:列假设 id | sec | scenario | question | points(json数组/分隔)
    points 列支持:
      - JSON 字符串:["得分点1","得分点2"]
      - 分隔字符串:得分点1|得分点2|得分点3
    """
    items = []
    headers = [cell_val(c).lower() for c in ws[1]]
    if 'id' not in headers:
        raise ValueError('案例表必须包含 id 列')
    sec_idx = headers.index('sec') if 'sec' in headers else None
    scen_idx = headers.index('scenario') if 'scenario' in headers else None
    q_idx = headers.index('question') if 'question' in headers else None
    pts_idx = headers.index('points') if 'points' in headers else None
    if q_idx is None or pts_idx is None:
        raise ValueError('案例表必须包含 question 和 points 列')
    for row in ws.iter_rows(min_row=2, values_only=False):
        rid = cell_val(row[headers.index('id')])
        if not rid:
            continue
        pts_raw = cell_val(row[pts_idx])
        pts = []
        if pts_raw:
            try:
                pts = json.loads(pts_raw)
            except json.JSONDecodeError:
                import re as _re
                pts = [p.strip() for p in _re.split(r'[|\n;；]', pts_raw) if p.strip()]
        item = {
            'id': rid,
            'question': cell_val(row[q_idx]),
            'points': pts,
        }
        if sec_idx is not None:
            sec = cell_val(row[sec_idx])
            if sec:
                item['sec'] = sec
        if scen_idx is not None:
            scen = cell_val(row[scen_idx])
            if scen:
                item['scenario'] = scen
        if item['question'] and item['points']:
            items.append(item)
    return items


# ---------- 文件读取入口 ----------

def read_file(path):
    """根据后缀分发: xlsx / csv / json"""
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xlsm'):
        return _read_xlsx(path)
    elif ext == '.csv':
        return _read_csv(path)
    elif ext == '.json':
        return _read_json(path)
    else:
        raise ValueError(f'不支持的文件格式: {ext} (仅支持 xlsx/csv/json)')


def _read_xlsx(path):
    """每个 Sheet 一类题型:sheet 名 = recite/quiz/exam/case"""
    wb = load_workbook(path, data_only=True)
    result = {'recite': [], 'quiz': [], 'exam': [], 'case': []}
    sheet_map = {'recite': xlsx_to_recite, 'quiz': xlsx_to_quiz,
                 'exam': lambda ws: xlsx_to_quiz(ws, 'exam'),
                 'case': xlsx_to_case}
    for sheet_name in wb.sheetnames:
        sn = sheet_name.strip().lower()
        if sn not in sheet_map:
            print(f'  ⚠ Sheet "{sheet_name}" 名称不在 recite/quiz/exam/case,跳过')
            continue
        ws = wb[sheet_name]
        items = sheet_map[sn](ws)
        result[sn].extend(items)
        print(f'  ✓ Sheet "{sheet_name}" → {len(items)} 条 {sn}')
    return result


def _read_csv(path):
    """单 CSV 文件,默认按 quiz 处理(可加 --type 参数覆盖)"""
    items = []
    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            row = {k.lower().strip(): v for k, v in row.items() if k}
            items.append(row)
    return items


def _read_json(path):
    """JSON 已是目标格式,直接读"""
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if isinstance(data, list):
        # 单题型列表,根据 --type 字段识别
        return data
    return data


# ---------- 主入口 ----------

def main():
    ap = argparse.ArgumentParser(description='二建备考App 题库转换器')
    ap.add_argument('input', help='输入文件路径(xlsx/csv/json)')
    ap.add_argument('--out', '-o', help='输出 JSON 文件路径(默认自动生成)')
    ap.add_argument('--type', '-t', choices=['recite','quiz','exam','case'],
                    help='强制指定题型(CSV/JSON 单文件时使用)')
    args = ap.parse_args()

    if not os.path.exists(args.input):
        print(f'❌ 文件不存在: {args.input}')
        sys.exit(1)

    print(f'📖 读取: {args.input}')
    data = read_file(args.input)

    # 统计
    if isinstance(data, list):
        # 单列表,按 --type 归类
        t = args.type or 'quiz'
        data = {t: data}
    # 确保 4 个键都存在
    for k in ('recite', 'quiz', 'exam', 'case'):
        data.setdefault(k, [])
    total = sum(len(v) for v in data.values())
    print(f'  recite:{len(data["recite"])} | quiz:{len(data["quiz"])} | '
          f'exam:{len(data["exam"])} | case:{len(data["case"])} (合计 {total})')

    # 输出
    if not args.out:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        base = os.path.splitext(os.path.basename(args.input))[0]
        args.out = os.path.join(os.path.dirname(args.input) or '.', f'{base}_{ts}.json')
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'✅ 已生成: {args.out}')


if __name__ == '__main__':
    main()